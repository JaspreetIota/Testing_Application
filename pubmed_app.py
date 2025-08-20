import streamlit as st
from Bio import Entrez
import pandas as pd
import re
from time import sleep
from io import BytesIO
import logging
import feedparser
import urllib.parse
from zipfile import ZipFile

# ---------- CONFIG ----------
Entrez.email = "your_email@example.com"  # Replace with your actual email
logging.basicConfig(level=logging.INFO)

# ---------- CONSTANTS ----------
PERSONAL_EMAIL_DOMAINS = ['gmail.com', 'yahoo.com', 'hotmail.com', 'outlook.com', 'aol.com']
BATCH_SIZE = 100

# ---------- FUNCTIONS ----------
def extract_email(text):
    match = re.search(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", text)
    return match.group(0).rstrip('.') if match else ""

def extract_university_name(affiliation_text):
    keywords = ['university', 'institute', 'college', 'school', 'center', 'centre', 'hospital']
    parts = [p.strip() for p in re.split(r'[;,]', affiliation_text.lower()) if p.strip()]
    matches = [p.title() for p in parts if any(k in p for k in keywords)]
    return matches[-1] if matches else ""

def format_mla(authors, title, journal, volume, issue, year, pages, doi):
    author_str = ", ".join(authors)
    return f"{author_str}. \"{title}.\" *{journal}*, vol. {volume}, no. {issue}, {year}, pp. {pages}. doi:{doi}"

def extract_doi(elocations):
    if isinstance(elocations, dict):
        elocations = [elocations]
    elif not isinstance(elocations, list):
        elocations = [elocations]
    for eloc in elocations:
        try:
            if hasattr(eloc, 'attributes') and eloc.attributes.get("EIdType") == "doi":
                return str(eloc)
        except Exception:
            continue
    return "N/A"

def get_google_news(query, max_articles=5):
    encoded_query = urllib.parse.quote(query)
    url = f'https://news.google.com/rss/search?q={encoded_query}'
    feed = feedparser.parse(url)
    news = []

    if not feed.entries:
        return []

    for entry in feed.entries[:max_articles]:
        news.append({
            'title': entry.title,
            'link': entry.link,
            'published': entry.published
        })
    return news

# ---------- STREAMLIT UI ----------
st.set_page_config(page_title="IOTA Tools", layout="wide")

menu = st.sidebar.selectbox("🔍 Select Tool", [
    "PubMed Article Extractor",
    "Google News Search",
    "Excel Splitter",
    "Yahoo Finance Company Lookup",
    "Excel Merger-Flatten Viewer"  # New Feature
])


# ==========================
# 🚀 PubMed Article Extractor
# ==========================
if menu == "PubMed Article Extractor":
    st.title("🔬 Auto Author/Article Extractor")

    search_term = st.text_input(
        "Enter your PubMed search term",
        '(Human Biology) AND ("united states"[Affiliation] OR USA[Affiliation]) AND (2022[Date - Publication])'
    )

    selected_countries = st.multiselect(
        "🌍 Select Countries (match in affiliation text)",
        options=[
            "USA", "United States", "United Kingdom", "Germany", "India", "Canada", "Australia",
            "France", "China", "Japan", "Brazil", "Italy", "Spain", "Netherlands", "Switzerland"
        ],
        default=["USA", "United States"]
    )

    retstart = st.number_input("Start from record number", min_value=0, value=0, step=100)
    retmax = st.number_input("How many records to fetch", min_value=10, max_value=10000, value=100, step=10)

    start_button = st.button("Fetch Articles")

    if not selected_countries:
        st.warning("⚠️ Please select at least one country to proceed.")

    if start_button and selected_countries:
        st.info(f"🔎 Searching PubMed (records {retstart} to {retstart + retmax - 1})...")
        try:
            search_handle = Entrez.esearch(
                db="pubmed",
                term=search_term,
                retstart=retstart,
                retmax=retmax
            )
            search_results = Entrez.read(search_handle)
        except Exception as e:
            st.error(f"❌ Failed to search PubMed: {e}")
            st.stop()

        pmids = search_results.get("IdList", [])
        if not pmids:
            st.error("No articles found for the given query and range.")
            st.stop()

        data = []
        progress = st.progress(0)

        for i, start in enumerate(range(0, len(pmids), BATCH_SIZE)):
            end = min(start + BATCH_SIZE, len(pmids))
            batch_pmids = pmids[start:end]

            try:
                fetch_handle = Entrez.efetch(db="pubmed", id=batch_pmids, rettype="xml")
                articles = Entrez.read(fetch_handle)
            except Exception as e:
                st.error(f"❌ Failed to fetch batch {start}-{end}: {e}")
                logging.exception("Fetch error")
                continue

            for article in articles['PubmedArticle']:
                try:
                    pmid = str(article['MedlineCitation']['PMID'])
                    article_data = article['MedlineCitation']['Article']
                    title = article_data.get("ArticleTitle", "No Title")
                    journal = article_data["Journal"]["Title"]
                    volume = article_data["Journal"]["JournalIssue"].get("Volume", "N/A")
                    issue = article_data["Journal"]["JournalIssue"].get("Issue", "N/A")
                    year = article_data["Journal"]["JournalIssue"]["PubDate"].get("Year", "N/A")
                    pages = article_data.get("Pagination", {}).get("MedlinePgn", "N/A")

                    doi = extract_doi(article_data.get("ELocationID", []))

                    for author in article_data.get("AuthorList", []):
                        if "ForeName" in author and "LastName" in author:
                            full_name = f"{author['LastName']}, {author['ForeName']}"

                            for aff in author.get("AffiliationInfo", []):
                                aff_text = aff.get("Affiliation", "")
                                email = extract_email(aff_text)
                                if not email or any(email.lower().endswith(f"@{domain}") for domain in PERSONAL_EMAIL_DOMAINS):
                                    continue

                                matched_country = next(
                                    (country for country in selected_countries
                                     if re.search(rf'\b{re.escape(country)}\b', aff_text, re.IGNORECASE)),
                                    None
                                )
                                if not matched_country:
                                    continue

                                university = extract_university_name(aff_text)
                                if not university:
                                    continue

                                mla = format_mla([full_name], title, journal, volume, issue, year, pages, doi)

                                data.append({
                                    "PMID": pmid,
                                    "Author": full_name,
                                    "Email": email,
                                    "Country": matched_country,
                                    "University": university,
                                    "Affiliation": aff_text,
                                    "MLA Citation": mla
                                })
                                break
                except Exception as e:
                    logging.exception("Article processing error")
                    st.warning(f"⚠️ Skipped an article due to error: {e}")
                    continue

            sleep(0.5)
            progress.progress((i + 1) / ((len(pmids) - 1) // BATCH_SIZE + 1))

        if data:
            df = pd.DataFrame(data)
            st.success(f"✅ Completed! {len(data)} valid entries extracted.")
            st.dataframe(df)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name="Results")
            output.seek(0)

            st.download_button(
                label="📁 Download Excel",
                data=output,
                file_name="pubmed_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("❌ No valid data extracted. Try refining your search.")

# =====================
# 📰 Google News Search
# =====================
elif menu == "Google News Search":
    st.title("📰 Google News Article Finder")

    query = st.text_input("Enter a topic, company, or keyword", "BYD Auto")
    max_articles = st.slider("Number of articles to display", 1, 20, 5)

    if st.button("Search News"):
        with st.spinner("Fetching news..."):
            news_results = get_google_news(query, max_articles=max_articles)

        if news_results:
            st.success(f"✅ Found {len(news_results)} articles.")
            for i, article in enumerate(news_results, 1):
                st.markdown(f"**{i}. [{article['title']}]({article['link']})**")
                st.markdown(f"*Published:* {article['published']}\n")
        else:
            st.warning("⚠️ No news articles found or feed could not be loaded.")

# =====================
# 📂 Excel Splitter
# =====================
elif menu == "Excel Splitter":
    st.title("📂 Excel Splitter - Split Large Excel File into Smaller Files")

    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

    entries_per_file = st.number_input(
        "Number of entries per file",
        min_value=100, max_value=10000, value=1500, step=100
    )

    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        total_rows = len(df)
        st.info(f"✅ Uploaded file has **{total_rows}** rows.")

        if st.button("🔨 Split Excel File"):
            with st.spinner("Splitting file..."):
                zip_buffer = BytesIO()
                with ZipFile(zip_buffer, "w") as zip_file:
                    num_files = (len(df) // entries_per_file) + (1 if len(df) % entries_per_file != 0 else 0)

                    for i in range(num_files):
                        start_row = i * entries_per_file
                        end_row = (i + 1) * entries_per_file
                        df_subset = df[start_row:end_row]

                        file_buffer = BytesIO()
                        df_subset.to_excel(file_buffer, index=False)
                        file_buffer.seek(0)

                        zip_file.writestr(f'split_part_{i + 1}.xlsx', file_buffer.read())

                zip_buffer.seek(0)

                st.download_button(
                    label="📦 Download ZIP of Split Files",
                    data=zip_buffer,
                    file_name="split_excel_files.zip",
                    mime="application/zip"
                )
# ===============================
# 🔎 Yahoo Finance Company Lookup
# ===============================
elif menu == "Yahoo Finance Company Lookup":
    st.title("🔎 Yahoo Finance Company Lookup")

    try:
        from yahooquery import search, Ticker
    except ImportError:
        st.error("❌ yahooquery is not installed. Please install it using `pip install yahooquery`")
        st.stop()

    def search_company_yahoo(name, limit=3):
        try:
            results = search(name)
            if not results or "quotes" not in results:
                return []
            companies = []
            for item in results['quotes']:
                if 'symbol' in item and 'shortname' in item:
                    companies.append({
                        "symbol": item['symbol'],
                        "name": item.get('shortname'),
                        "exchange": item.get('exchDisp'),
                        "type": item.get('typeDisp')
                    })
                    if len(companies) >= limit:
                        break
            return companies
        except Exception as e:
            return {"error": str(e)}

    def get_company_info(symbol):
        try:
            ticker = Ticker(symbol)
            profile = ticker.asset_profile
            info = profile.get(symbol, {})
            return info
        except Exception as e:
            return {"error": str(e)}

    st.markdown("Search for a company and fetch detailed metadata from Yahoo Finance.")
    company_input = st.text_input("Enter Company Name", value="Apple")

    if st.button("Search Company"):
        if not company_input.strip():
            st.warning("Please enter a company name.")
            st.stop()

        with st.spinner("Searching Yahoo Finance..."):
            matches = search_company_yahoo(company_input)

        if isinstance(matches, dict) and "error" in matches:
            st.error(f"Error: {matches['error']}")
        elif not matches:
            st.warning("No matches found.")
        else:
            st.success(f"Found {len(matches)} match(es). Showing top result.")

            top_match = matches[0]
            st.markdown(f"**Top Match:** `{top_match['name']}`")
            st.markdown(f"- **Symbol:** {top_match['symbol']}")
            st.markdown(f"- **Exchange:** {top_match['exchange']}")
            st.markdown(f"- **Type:** {top_match['type']}")

            with st.spinner("Fetching company details..."):
                details = get_company_info(top_match['symbol'])

            if isinstance(details, dict) and "error" in details:
                st.error(f"Error: {details['error']}")
            elif details:
                st.subheader("📊 Company Details")
                for key, value in details.items():
                    st.markdown(f"- **{key}:** {value}")
            else:
                st.warning("No detailed data available for this symbol.")
# ===============================
# 🧾 Excel Merger-Flatten Viewer
# ===============================
elif menu == "Excel Merger-Flatten Viewer":
    st.title("🧾 Flatten Excel with Merged Cells")

    uploaded_file = st.file_uploader("Upload an Excel file with merged cells", type=["xlsx"])

    if uploaded_file:
        from openpyxl import load_workbook

        with st.spinner("Processing..."):
            wb = load_workbook(uploaded_file)
            ws = wb.active

            # Step 1: Read values into 2D list
            data = [[cell.value for cell in row] for row in ws.iter_rows()]

            # Step 2: Fill in merged cell ranges
            for merged_range in ws.merged_cells.ranges:
                min_row, min_col, max_row, max_col = (
                    merged_range.min_row,
                    merged_range.min_col,
                    merged_range.max_row,
                    merged_range.max_col
                )
                top_left_value = ws.cell(row=min_row, column=min_col).value

                for row in range(min_row - 1, max_row):
                    for col in range(min_col - 1, max_col):
                        if row != min_row - 1 or col != min_col - 1:
                            data[row][col] = top_left_value

            # Step 3: Convert to DataFrame and display
            df = pd.DataFrame(data)
            st.success("✅ File processed. Here's a preview:")
            st.dataframe(df)

            # Step 4: Download cleaned Excel
            output_buffer = BytesIO()
            df.to_excel(output_buffer, index=False, header=False)
            output_buffer.seek(0)

            st.download_button(
                label="📥 Download Flattened Excel",
                data=output_buffer,
                file_name="flattened_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


